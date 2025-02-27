import pygame
import random
import sys
from openpyxl import load_workbook

pygame.init()
WIDTH, HEIGHT = 800, 600
FPS = 60
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)
GREEN = (17, 95, 17)
RED = (255, 0, 0)
screen = pygame.display.set_mode((WIDTH, HEIGHT))
pygame.display.set_caption("Dungeon Card Game")
clock = pygame.time.Clock()
font = pygame.font.Font(None, 24)
large_font = pygame.font.Font(None, 48)
try:
    dungeon_bg1 = pygame.image.load("fon2.png")
    dungeon_bg2 = pygame.image.load("fon.png")
    table_image = pygame.image.load("fon_osn.png")
except pygame.error:
    print("Ошибка загрузки изображений!")
    sys.exit()
def load_cards():
    try:
        workbook = load_workbook("Карты для манчкина.xlsx")
        sheet = workbook.active
        cards = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            card = {
                "name": row[0],
                "type": row[1],
                "bonus": row[2] if row[2] else 0
            }
            cards.append(card)
        return cards
    except Exception as e:
        print(f"Ошибка загрузки карт: {e}")
        sys.exit()

class Player:
    def __init__(self, name):
        self.name = name
        self.health = 100
        self.level = 1
        self.hand = []

    def draw_card(self, deck):
        if deck:
            card = random.choice(deck)
            self.hand.append(card)
            deck.remove(card)
            return card
        return None

    def attack(self, opponent):
        damage = random.randint(5, 20)
        opponent.health -= damage
        return damage

    def heal(self):
        heal_amount = random.randint(10, 30)
        self.health += heal_amount
        return heal_amount

    def use_item(self, item_name):
        for card in self.hand:
            if card["type"] == "Item" and card["name"] == item_name:
                self.level += card.get("bonus", 0)
                self.hand.remove(card)
                return True
        return False

class Game:
    def __init__(self):
        self.players = [Player("Игрок 1"), Player("Игрок 2")]
        self.current_player = 0
        self.deck = load_cards()
        self.running = True

    def switch_player(self):
        self.current_player = (self.current_player + 1) % 2

    def get_current_player(self):
        return self.players[self.current_player]

    def get_opponent(self):
        return self.players[(self.current_player + 1) % 2]

def draw_text(text, x, y, font, color=BLACK):
    text_surface = font.render(text, True, color)
    screen.blit(text_surface, (x, y))

def show_rules():
    screen.blit(table_image, (0, 0))
    rules = [
        "Правила игры:",
        "Настольная игра Манчкин!",
        "Она отправляет вас в мир сражений, дружеской поддержки и коварства на пути к победе!",
        " Это свежий взгляд на ролевые игры, наполненные многообразием ходов.",
        "Цель игры дойти до 10 уровня.",
        "Повысить его можно за счет специальных карт, побед над монстрами или продажи шмоток.",
        "Манчкин полностью погружает вас в абсолютно новый мир со своими правилами и языком!",
        "Важно помнить, что 10 уровень можно получить, только победив монстра и никак иначе",
        "В самом начале хода перед вами многообразие вариантов действия:",
        "можете выбрать игрока, атаковать и сражаться, а главное, дойти до 10 уровня и выиграть.",
        "Также ты сможешь одержать победу, растратив все очки здоровья своего противника.",
        "Нажмите любую клавишу для возврата."
    ]
    for i, rule in enumerate(rules):
        draw_text(rule, 20, 20 + i * 30, font, BLACK)
    pygame.display.flip()
    wait_for_key()

def wait_for_key():
    while True:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.KEYDOWN:
                return

def show_item_menu(player):
    screen.blit(table_image, (0, 0))
    draw_text("Выберите предмет для использования:", 20, 20, font, BLACK)
    items = [
        "Баклер бахвала",
        "Бензопила кровавого расчленения",
        "Дуб джентельменов",
        "Кинжал измены",
        "Крыса на палочке",
        "Лучок с ленточками",
        "Меч широты взглядов",
        "Коленоотбойный молот",
        "Огромная каменюга",
        "Одиннадцатифутовый шест",
        "Палица остроты",
        "Меч коварного ублюдка",
        "Посох напалма",
        "Рапира такнечестности",
        "Совсехсторонний щит",
        "Сыротёрка-умиротворения",
        "Чарующая дуда",
        "Швейцарская армейская алебарда"
    ]
    if not items:
        draw_text("У вас нет предметов!", 20, 50, font, RED)
        pygame.display.flip()
        pygame.time.delay(2000)
        return None
    for i, item in enumerate(items):
        draw_text(f"{i + 1}. {item}", 20, 50 + i * 30, font, BLACK)
    pygame.display.flip()
    while True:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.KEYDOWN:
                if event.key >= pygame.K_1 and event.key <= pygame.K_9:
                    index = event.key - pygame.K_1
                    if 0 <= index < len(items):
                        return items[index]

def monster_battle(player, opponent):
    screen.blit(table_image, (0, 0))
    draw_text("Сражение с монстром:", 20, 20, large_font, BLACK)
    draw_text("1. Сразиться", 20, 100, font, BLACK)
    draw_text("2. Предметы", 20, 130, font, BLACK)
    draw_text("3. Сбежать", 20, 160, font, BLACK)
    pygame.display.flip()

    while True:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_1:
                    damage = random.randint(10, 30)
                    opponent.health -= damage
                    draw_text(f"Вы нанесли {damage} урона!", 20, 200, font, RED)
                    pygame.display.flip()
                    pygame.time.delay(2000)
                    return True
                elif event.key == pygame.K_2:
                    item_name = show_item_menu(player)
                    if item_name:
                        player.use_item(item_name)
                        draw_text(f"Вы использовали {item_name}!", 250, 200, font, GREEN)
                        pygame.display.flip()
                        pygame.time.delay(2000)
                        return True
                    else:
                        draw_text("Вы не выбрали предмет и проиграли!", 250, 200, font, RED)
                        pygame.display.flip()
                        pygame.time.delay(2000)
                        return False
                elif event.key == pygame.K_3:  # Сбежать
                    if random.random() < 0.5:
                        draw_text("Вы успешно сбежали!", 250, 200, font, GREEN)
                        pygame.display.flip()
                        pygame.time.delay(2000)
                        return False
                    else:
                        draw_text("Вы не смогли сбежать и проиграли!", 20, 200, font, RED)
                        pygame.display.flip()
                        pygame.time.delay(2000)
                        return False

def draw_menu():
    screen.blit(dungeon_bg1, (0, 0))
    draw_text("MANCHKIN", WIDTH // 2 - 150, HEIGHT // 4, pygame.font.Font(None, 100), (20, 180, 100))
    draw_text("1. Начать игру", WIDTH // 2 - 100, HEIGHT // 2, pygame.font.Font(None, 48), BLACK)
    draw_text("2. Правила", WIDTH // 2 - 100, HEIGHT // 2 + 50, pygame.font.Font(None, 48), BLACK)
    draw_text("3. Выход", WIDTH // 2 - 100, HEIGHT // 2 + 100, pygame.font.Font(None, 48), BLACK)
    pygame.display.flip()

def pause_menu():
    paused = True
    while paused:
        screen.fill(WHITE)
        draw_text("ПАУЗА", WIDTH // 2 - 50, HEIGHT // 4, large_font, RED)
        draw_text("1. Продолжить", WIDTH // 2 - 100, HEIGHT // 2, font, BLACK)
        draw_text("2. Выйти в главное меню", WIDTH // 2 - 100, HEIGHT // 2 + 50, font, BLACK)
        pygame.display.flip()
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_1:
                    paused = False
                elif event.key == pygame.K_2:
                    main_menu()

def game_loop():
    game = Game()
    running = True
    while running:
        screen.blit(dungeon_bg1, (0, 0))
        player = game.get_current_player()
        opponent = game.get_opponent()
        draw_text(f"{player.name} - Уровень: {player.level}, Здоровье: {player.health}", 20, 20, font)
        draw_text(f"{opponent.name} - Уровень: {opponent.level}, Здоровье: {opponent.health}", 20, 50, font)
        draw_text("1. Выбрать монстра", 20, 100, font)
        draw_text("2. Атаковать", 20, 130, font)
        draw_text("3. Лечиться", 20, 160, font)
        draw_text("4. Сражение с монстром", 20, 190, font)
        draw_text("P. Пауза", 20, 220, font)
        pygame.display.flip()
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_p:
                    pause_menu()
                if event.key == pygame.K_1:
                    card = player.draw_card(game.deck)
                    if card:
                        draw_text(f"Вы выбрали монстра: {card['name']}", 20, 250, font, GREEN)
                    else:
                        draw_text("Колода пуста!", 20, 250, font, RED)
                    pygame.display.flip()
                    pygame.time.delay(1000)

                elif event.key == pygame.K_2:
                    damage = player.attack(opponent)
                    draw_text(f"Вы нанесли {damage} урона!", 20, 250, font, RED)
                    if opponent.health <= 0:
                        draw_text(f"{opponent.name} проиграл!", WIDTH // 2 - 100, HEIGHT // 2, large_font, RED)
                        pygame.display.flip()
                        pygame.time.delay(2000)
                        running = False
                    pygame.display.flip()
                    pygame.time.delay(1000)

                elif event.key == pygame.K_3:
                    heal_amount = player.heal()
                    draw_text(f"Вы восстановили {heal_amount} здоровья!", 20, 250, font, GREEN)
                    pygame.display.flip()
                    pygame.time.delay(1000)

                elif event.key == pygame.K_4:
                    if monster_battle(player, opponent):
                        player.level += 1
                        draw_text(f"Вы победили! Ваш уровень: {player.level}", 250, 250, font, GREEN)
                        if player.level >= 10:
                            draw_text(f"{player.name} выиграл игру!", WIDTH // 2 - 100, HEIGHT // 2, large_font, GREEN)
                            pygame.display.flip()
                            pygame.time.delay(2000)
                            running = False
                    else:
                        draw_text("Вы проиграли битву!", 250, 250, font, RED)
                    pygame.display.flip()
                    pygame.time.delay(1000)

                game.switch_player()

        clock.tick(FPS)

def main_menu():
    while True:
        draw_menu()
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_1:
                    game_loop()
                elif event.key == pygame.K_2:
                    show_rules()
                elif event.key == pygame.K_3:
                    pygame.quit()
                    sys.exit()

if __name__ == "__main__":
    main_menu()